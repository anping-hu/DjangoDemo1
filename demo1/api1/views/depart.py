from django.shortcuts import render, HttpResponse, redirect
from openpyxl import load_workbook
from api1 import models
from api1.utils.pagination import Pagination


def depart_list(request):
    """ 部门列表 """
    query = models.Department.objects.all()  # 从数据库中拿到所有部门表的数据，再传入html里面
    # 分页
    page_object = Pagination(request, query)
    context = {
        "query": page_object.page_queryset,  # 分完页的数据
        "page_string": page_object.html()  # 生成页码
    }

    return render(request, 'depart_list.html', context)  # 传变量到html中是字典类型


def depart_add(request):
    """ 添加部门函数 """
    print('请求方式是：', request.method)
    if request.method == "GET":
        return render(request, 'depart_add.html')
    # 如果是post请求过来，即点提交按钮，获取请求中的内容
    title = request.POST.get("title")
    print('title是：', title)
    if title == '':
        return render(request, 'depart_add.html')
    # 保存到数据库中
    models.Department.objects.create(title=title)
    # 重定向回部门页面
    return redirect("/depart/list/")


def depart_delete(request):
    """ 删除部门函数 """
    id = request.GET.get("id")  # 得到get请求里面的id
    print('删除部门函数id:', id)
    models.Department.objects.filter(id=id).delete()  # 根据部门id删除数据库中对应的部门
    return redirect("/depart/list/")  # 重定向


def depart_edit(request, nid):
    """ 编辑部门 """
    if request.method == "GET":
        d_object = models.Department.objects.filter(id=nid).first()
        return render(request, 'depart_edit.html', {"d_object": d_object})
    # 如果是POST请求
    Utitle = request.POST.get("title")
    models.Department.objects.filter(id=nid).update(title=Utitle)
    return redirect("/depart/list/")  # 重定向


def depart_multi(request):
    """ 批量删除（Excel文件）"""

    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")

    if file_object:
        # 2.对象传递给openpyxl，由openpyxl读取文件的内容
        wb = load_workbook(file_object)
        sheet = wb.worksheets[0]

        # 3.循环获取每一行数据
        for row in sheet.iter_rows(min_row=2):  # min_row=2表示从表格的第二行读取
            text = row[0].value  # row[0] 表示每一行的第一列
            # 判断新增部门在数据库中是否存在，不存在添加
            exists = models.Department.objects.filter(title=text).exists()
            if not exists:
                models.Department.objects.create(title=text)

        return redirect('/depart/list/')
    else:
        return HttpResponse('批量上传未选择文件')
